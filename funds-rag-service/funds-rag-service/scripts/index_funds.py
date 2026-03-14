#!/usr/bin/env python3
"""
Index funds from JSONL (primary), CSV, or Excel (.xls) into PostgreSQL with pgvector embeddings.

Usage:
    python -m scripts.index_funds --jsonl /app/funds_data/funds_clean.jsonl
    python -m scripts.index_funds --xls /app/funds_data/funds_clean.xls
    python -m scripts.index_funds --csv data/funds.csv   # legacy
"""
import json
import csv
import argparse
import sys
import time
from urllib.parse import urlparse

from sqlalchemy import text
from sqlalchemy.orm import Session

# Ensure project root is on path
sys.path.insert(0, ".")

from app.config import settings
from app.models import Base, Fund
from app.embeddings import get_embeddings_batch, build_fund_text, build_fund_text_from_jsonl
from app.fund_parsers import normalize_stages, parse_check_size_to_usd
from app.database import sync_engine


def create_tables():
    """Create pgvector extension, tables, columns and performance indexes."""
    with sync_engine.connect() as conn:
        # pgvector extension
        conn.execute(text("CREATE EXTENSION IF NOT EXISTS vector"))
        conn.commit()

    # Base tables (SQLAlchemy models)
    Base.metadata.create_all(sync_engine)

    # Migrations: new columns + indexes for fast vector + structured filters
    with sync_engine.connect() as conn:
        for stmt in (
            # New columns (idempotent)
            "ALTER TABLE funds ADD COLUMN IF NOT EXISTS stage TEXT[]",
            "ALTER TABLE funds ADD COLUMN IF NOT EXISTS check_min_usd NUMERIC(14,2)",
            "ALTER TABLE funds ADD COLUMN IF NOT EXISTS check_max_usd NUMERIC(14,2)",
            "ALTER TABLE funds ADD COLUMN IF NOT EXISTS check_size_text TEXT",
            # Vector index (ivfflat) on embedding
            "CREATE INDEX IF NOT EXISTS funds_embedding_ivfflat "
            "ON funds USING ivfflat (embedding vector_cosine_ops) WITH (lists = 100)",
            # GIN index on stage array for fast stage filters
            "CREATE INDEX IF NOT EXISTS funds_stage_gin "
            "ON funds USING GIN (stage)",
            # B-tree index on check range for range filters
            "CREATE INDEX IF NOT EXISTS funds_check_range "
            "ON funds (check_min_usd, check_max_usd)",
        ):
            conn.execute(text(stmt))
        # Update planner statistics so it uses the new indexes
        conn.execute(text("ANALYZE funds"))
        conn.commit()

    print("[OK] Tables and indexes created")


# ─── JSONL (funds_clean.jsonl / investors_rag-style) ───────────────────────

def load_jsonl(path: str) -> list[dict]:
    """Load JSONL into list of dicts."""
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    print(f"[OK] Loaded {len(rows)} rows from {path}")
    return rows


def jsonl_row_to_fund_data(row: dict) -> dict:
    """Map funds JSONL fields (e.g. funds_clean.jsonl) to Fund model (no embedding)."""
    website = (row.get("website") or "").strip()
    domain = None
    if website:
        try:
            domain = urlparse(website).netloc or website
        except Exception:
            domain = website

    stages_raw = row.get("stages") or ""
    stage_list = normalize_stages(stages_raw)
    check_min, check_max = parse_check_size_to_usd(row.get("check_size"))

    return {
        "sno": int(row["id"]) if row.get("id") and str(row["id"]).isdigit() else None,
        "investor_name": (row.get("name") or "").strip() or "Unknown",
        "domain_name": domain,
        "overview": (row.get("description") or "").strip()[:5000] if row.get("description") else None,
        "founded_year": str(row.get("founded_year") or "")[:20] if row.get("founded_year") else None,
        "country": (row.get("hq_country") or "").strip()[:500] or None,
        "state": None,
        "city": (row.get("hq_city") or "").strip()[:500] or None,
        "description": (row.get("description") or "").strip()[:10000] if row.get("description") else None,
        "investor_type": (row.get("type") or "").strip()[:200] or None,
        "practice_areas": (row.get("fund_model") or "").strip()[:500] or None,
        "feed_name": None,
        "business_models": (row.get("sectors") or "").strip()[:1000] or None,
        "investment_score": None,
        "website": website[:1000] if website else None,
        "linkedin": (row.get("linkedin") or "").strip()[:1000] or None,
        "twitter": None,
        "stage": stage_list if stage_list else None,
        "check_min_usd": check_min,
        "check_max_usd": check_max,
        "check_size_text": (row.get("check_size") or "").strip() or None,
    }


def index_jsonl(jsonl_path: str, batch_size: int = 50):
    """Index from funds JSONL (e.g. funds_clean.jsonl); use 'text' field for embeddings."""
    create_tables()
    rows = load_jsonl(jsonl_path)
    if not rows:
        print("[WARN] No rows in JSONL")
        return

    texts = [build_fund_text_from_jsonl(row) for row in rows]
    print(f"[..] Generating embeddings for {len(texts)} funds (batch_size={batch_size})...")
    t0 = time.time()
    embeddings = get_embeddings_batch(texts, batch_size=batch_size)
    elapsed = time.time() - t0
    print(f"[OK] Embeddings generated in {elapsed:.1f}s")

    session = Session(sync_engine)
    try:
        session.execute(text("TRUNCATE TABLE funds RESTART IDENTITY"))
        session.commit()
        print("[..] Inserting funds into database...")

        for i, (row, emb) in enumerate(zip(rows, embeddings)):
            fund_data = jsonl_row_to_fund_data(row)
            fund = Fund(**fund_data, embedding=emb)
            session.add(fund)
            if (i + 1) % 200 == 0:
                session.commit()
                print(f"  [{i + 1}/{len(rows)}] committed")

        session.commit()
        print(f"[OK] All {len(rows)} funds indexed successfully")
    except Exception as e:
        session.rollback()
        print(f"[ERROR] {e}")
        raise
    finally:
        session.close()


# ─── CSV (legacy) ──────────────────────────────────────────────────────────

COL_MAP = {
    "SNo.": "sno",
    "Investor Name": "investor_name",
    "Domain Name": "domain_name",
    "Overview": "overview",
    "Founded Year": "founded_year",
    "Country": "country",
    "State": "state",
    "City": "city",
    "Description": "description",
    "Investor Type": "investor_type",
    "Practice Areas": "practice_areas",
    "Feed Name": "feed_name",
    "Business Models": "business_models",
    "Investment Score": "investment_score",
    "Website": "website",
    "LinkedIn": "linkedin",
    "Twitter": "twitter",
}


def _row_to_snake_case(row: dict) -> dict:
    return {model_col: (row.get(csv_col) or "").strip() for csv_col, model_col in COL_MAP.items()}


def load_csv(csv_path: str) -> list[dict]:
    rows = []
    with open(csv_path, "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            rows.append(row)
    print(f"[OK] Loaded {len(rows)} rows from {csv_path}")
    return rows


# ─── Excel (.xls) ───────────────────────────────────────────────────────────

# Map Excel header (strip, lower) -> JSONL-style key for jsonl_row_to_fund_data
XLS_HEADER_TO_JSONL = {
    "investor name": "name", "name": "name",
    "country": "hq_country", "hq country": "hq_country", "hq_country": "hq_country",
    "city": "hq_city", "hq city": "hq_city", "hq_city": "hq_city",
    "sectors": "sectors", "sector": "sectors",
    "stages": "stages", "stage": "stages",
    "check size": "check_size", "check_size": "check_size",
    "description": "description", "overview": "overview",
    "website": "website", "linkedin": "linkedin", "linkedin url": "linkedin",
    "founded year": "founded_year", "founded_year": "founded_year",
    "type": "type", "investor type": "type",
    "fund model": "fund_model", "fund_model": "fund_model",
    "id": "id", "sno": "id", "sno.": "id",
    "text": "text",  # pre-built text for embedding
    "domain name": "domain_name", "domain_name": "domain_name",
}


def _xls_row_to_jsonl_like(row: dict, original_header_to_jsonl: dict) -> dict:
    """Convert one XLS row (keys = original sheet headers) to JSONL-like dict."""
    out = {}
    for orig_header, value in row.items():
        jsonl_key = original_header_to_jsonl.get((orig_header or "").strip())
        if jsonl_key is None:
            continue
        if value is None or (isinstance(value, float) and value != value):  # NaN
            val = ""
        else:
            val = str(value).strip() if value else ""
        if jsonl_key == "overview":
            out.setdefault("description", val)
        else:
            out[jsonl_key] = val
    return out


def _jsonl_to_legacy_row(row: dict) -> dict:
    """Map JSONL-like keys to legacy build_fund_text keys (for XLS path)."""
    return {
        "investor_name": row.get("name") or "",
        "country": row.get("hq_country") or "",
        "city": row.get("hq_city") or "",
        "investor_type": row.get("type") or "",
        "description": row.get("description") or row.get("overview") or "",
        "overview": row.get("description") or row.get("overview") or "",
        "practice_areas": row.get("fund_model") or "",
        "business_models": row.get("sectors") or "",
        "founded_year": str(row.get("founded_year") or ""),
        "website": row.get("website") or "",
        "linkedin": row.get("linkedin") or "",
    }


def load_xls(xls_path: str) -> list[dict]:
    """Load first sheet of .xls or .xlsx; first row = headers. Returns list of dicts (JSONL-like keys)."""
    path_lower = xls_path.lower()
    if path_lower.endswith(".xlsx"):
        return _load_xlsx(xls_path)
    import xlrd
    with xlrd.open_workbook(xls_path) as wb:
        sh = wb.sheet_by_index(0)
        headers = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
        orig_to_jsonl = _build_header_map(headers)
        rows = []
        for r in range(1, sh.nrows):
            row = {headers[c]: sh.cell_value(r, c) for c in range(sh.ncols)}
            rows.append(_xls_row_to_jsonl_like(row, orig_to_jsonl))
    print(f"[OK] Loaded {len(rows)} rows from {xls_path}")
    return rows


def _build_header_map(headers: list) -> dict:
    """Map original header -> jsonl key for known columns."""
    orig_to_jsonl = {}
    for h in headers:
        k = (h or "").lower().strip()
        if k in XLS_HEADER_TO_JSONL:
            orig_to_jsonl[h] = XLS_HEADER_TO_JSONL[k]
    return orig_to_jsonl


def _load_xlsx(xlsx_path: str) -> list[dict]:
    """Load first sheet of .xlsx using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sh = wb.active
    headers = [str(sh.cell(1, c).value or "").strip() for c in range(1, sh.max_column + 1)]
    orig_to_jsonl = _build_header_map(headers)
    rows = []
    for r in range(2, sh.max_row + 1):
        row = {headers[c]: sh.cell(r, c + 1).value for c in range(len(headers))}
        rows.append(_xls_row_to_jsonl_like(row, orig_to_jsonl))
    wb.close()
    print(f"[OK] Loaded {len(rows)} rows from {xlsx_path}")
    return rows


def index_xls(xls_path: str, batch_size: int = 50):
    """Index from Excel .xls; map columns to JSONL-style and use same pipeline as JSONL."""
    create_tables()
    rows = load_xls(xls_path)
    # Build text for embedding (use 'text' column if present, else from fields)
    texts = []
    for row in rows:
        if row.get("text"):
            texts.append((row.get("text") or "").strip())
        else:
            texts.append(build_fund_text(_jsonl_to_legacy_row(row)))
    # Fallback if no text at all
    texts = [t or "Unknown fund" for t in texts]
    print(f"[..] Generating embeddings for {len(texts)} funds (batch_size={batch_size})...")
    t0 = time.time()
    embeddings = get_embeddings_batch(texts, batch_size=batch_size)
    print(f"[OK] Embeddings generated in {time.time() - t0:.1f}s")

    session = Session(sync_engine)
    try:
        session.execute(text("TRUNCATE TABLE funds RESTART IDENTITY"))
        session.commit()
        for i, (row, emb) in enumerate(zip(rows, embeddings)):
            fund_data = jsonl_row_to_fund_data(row)
            if not fund_data.get("investor_name"):
                fund_data["investor_name"] = "Unknown"
            session.add(Fund(**fund_data, embedding=emb))
            if (i + 1) % 200 == 0:
                session.commit()
                print(f"  [{i + 1}/{len(rows)}] committed")
        session.commit()
        print(f"[OK] All {len(rows)} funds indexed from XLS")
    except Exception as e:
        session.rollback()
        raise
    finally:
        session.close()


def index_csv(csv_path: str, batch_size: int = 50):
    """Legacy: index from CSV."""
    create_tables()
    rows = load_csv(csv_path)
    texts = [build_fund_text(_row_to_snake_case(row)) for row in rows]
    print(f"[..] Generating embeddings for {len(texts)} funds (batch_size={batch_size})...")
    t0 = time.time()
    embeddings = get_embeddings_batch(texts, batch_size=batch_size)
    print(f"[OK] Embeddings generated in {time.time() - t0:.1f}s")

    session = Session(sync_engine)
    try:
        session.execute(text("TRUNCATE TABLE funds RESTART IDENTITY"))
        session.commit()
        for i, (row, emb) in enumerate(zip(rows, embeddings)):
            fund_data = {}
            for csv_col, model_col in COL_MAP.items():
                val = row.get(csv_col, "").strip()
                if model_col == "investment_score":
                    fund_data[model_col] = float(val) if val else None
                elif model_col == "sno":
                    fund_data[model_col] = int(val) if val else None
                else:
                    fund_data[model_col] = val if val else None
            session.add(Fund(**fund_data, embedding=emb))
            if (i + 1) % 200 == 0:
                session.commit()
                print(f"  [{i + 1}/{len(rows)}] committed")
        session.commit()
        print(f"[OK] All {len(rows)} funds indexed successfully")
    except Exception as e:
        session.rollback()
        raise
    finally:
        session.close()


# ─── CLI ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Index funds (JSONL, Excel .xls, or CSV) into pgvector")
    parser.add_argument("--jsonl", default=settings.funds_jsonl_path, help="Path to funds JSONL (e.g. funds_clean.jsonl)")
    parser.add_argument("--xls", help="Path to Excel .xls (e.g. funds_clean.xls)")
    parser.add_argument("--csv", help="Path to CSV (legacy)")
    parser.add_argument("--batch-size", type=int, default=50)
    args = parser.parse_args()

    if args.xls:
        index_xls(args.xls, args.batch_size)
    elif args.csv:
        index_csv(args.csv, args.batch_size)
    else:
        index_jsonl(args.jsonl, args.batch_size)


if __name__ == "__main__":
    main()
